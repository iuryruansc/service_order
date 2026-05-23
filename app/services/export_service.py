import io
import openpyxl
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from sqlalchemy.orm import Session
from app.repositories.service_order_repository import get_service_orders

def export_to_excel(db: Session, exported_by: str) -> io.BytesIO:
    orders = get_service_orders(db)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Service Orders"

    headers = ["ID", "Título", "Descrição", "Status", "Prioridade", "Cliente", "Responsável", "Exportado por", "Data de Criação"]
    sheet.append(headers)

    for order in orders:
        sheet.append([
            order.id,
            order.title,
            order.description,
            order.status,
            order.priority,
            order.client_id,
            order.responsible_user_id,
            exported_by,
            order.created_at.strftime("%d/%m/%Y %H:%M"),
        ])

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 4

    file_content = io.BytesIO()
    workbook.save(file_content)
    file_content.seek(0)

    return file_content

def export_to_pdf(db: Session, exported_by: str) -> io.BytesIO:
    orders = get_service_orders(db)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)

    pdf.set_font("Helvetica", style="B", size=14)
    pdf.cell(0, 10, "Relatório de Ordens de Serviço", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", size=8)
    pdf.cell(0, 6, f"Exportado por: {exported_by}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", style="B", size=8)
    headers = ["ID", "Título", "Status", "Prioridade", "Cliente", "Responsável", "Data"]
    widths = [10, 50, 25, 25, 25, 25, 25]

    for header, width in zip(headers, widths):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", size=8)
    for order in orders:
        values = [
            str(order.id),
            order.title,
            order.status,
            order.priority,
            str(order.client_id),
            str(order.responsible_user_id),
            order.created_at.strftime("%d/%m/%Y"),
        ]
        for value, width in zip(values, widths):
            pdf.cell(width, 7, value[:20], border=1)
        pdf.ln()

    file_content = io.BytesIO()
    pdf.output(file_content)
    file_content.seek(0)

    return file_content